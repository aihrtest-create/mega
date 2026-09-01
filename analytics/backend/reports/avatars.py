import requests
import pandas as pd
from datetime import datetime, timedelta
from backend.reports.common import GRAFANA_URL, DATASOURCE_UID, get_headers, get_time_boundaries, normalize_park_name
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def generate_avatars(date_val: str, period_type: str, selected_parks: list, output_path: str):
    start_date, stop_date, _ = get_time_boundaries(date_val, period_type)
    
    flux_query = f'''
    from(bucket: "Analytics_AvatarBD")
      |> range(start: {start_date}, stop: {stop_date})
      |> filter(fn: (r) => r["_measurement"] == "AvatarLinkedInHome")
      |> group(columns: ["park"])
      |> aggregateWindow(every: 1h, fn: count, createEmpty: false)
      |> yield(name: "avatars_by_hour")
    '''

    payload = {
        "queries": [{"refId": "A", "datasource": {"type": "influxdb", "uid": DATASOURCE_UID}, "query": flux_query, "hide": False}],
        "from": "0", "to": "9999999999999"
    }

    response = requests.post(f"{GRAFANA_URL}/api/ds/query", headers=get_headers(), json=payload, timeout=30)

    response.raise_for_status()
    
    frames = response.json().get("results", {}).get("A", {}).get("frames", [])
    results = []
    
    for frame in frames:
        park_name = "Unknown"
        for field in frame.get("schema", {}).get("fields", []):
            if "park" in field.get("labels", {}):
                park_name = field["labels"]["park"]
                break
        
        park_name = normalize_park_name(park_name)

        
        vals = frame.get("data", {}).get("values", [])
        if len(vals) >= 2:
            for ts, val in zip(vals[0], vals[1]):
                if isinstance(ts, (int, float)):
                    dt = datetime.fromtimestamp(ts / 1000)
                else:
                    dt = pd.to_datetime(ts)
                dt = dt + timedelta(hours=3)
                results.append({"Дата": dt.strftime("%Y-%m-%d"), "Парк": park_name, "Привязано аватаров": val})

    if not results:
        raise Exception("Нет данных за выбранный период")

    df = pd.DataFrame(results)
    
    # Filter by user selected parks
    df = df[df["Парк"].isin(selected_parks)]
    if df.empty:
        raise Exception("Нет данных для выбранных парков")
        
    pivot_df = df.pivot_table(index="Дата", columns="Парк", values="Привязано аватаров", aggfunc="sum").fillna(0).astype(int)
    
    # Сортируем колонки по алфавиту для красоты (только выбранные)
    pivot_df = pivot_df.reindex(sorted(pivot_df.columns, key=lambda x: str(x).lower()), axis=1)
    
    # Фильтр по периоду (чтобы убрать лишние дни, возникшие из-за смещения таймзон)
    if period_type == "month":
        pivot_df = pivot_df[pivot_df.index.str.startswith(date_val)]
    elif period_type == "day":
        pivot_df = pivot_df[pivot_df.index == date_val]
    elif period_type == "year":
        pivot_df = pivot_df[pivot_df.index.str.startswith(date_val)]
    elif period_type == "custom":
        start_str, stop_str = date_val.split("_to_")
        pivot_df = pivot_df[(pivot_df.index >= start_str) & (pivot_df.index <= stop_str)]
    
    pivot_df["ИТОГО"] = pivot_df.sum(axis=1)
    totals = pivot_df.sum(axis=0)
    totals.name = "ИТОГО по парку"
    pivot_df = pd.concat([pivot_df, totals.to_frame().T])
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pivot_df.to_excel(writer, sheet_name=date_val)
        ws = writer.sheets[date_val]
        
        last_data_row = ws.max_row
        last_col = ws.max_column
        
        orange = PatternFill("solid", fgColor="FF6B00")
        for cell in ws[1]:
            cell.fill = orange; cell.font = Font(bold=True, color="FFFFFF"); cell.alignment = Alignment(horizontal="center")
            
        gray = PatternFill("solid", fgColor="E0E0E0")
        for cell in ws[last_data_row]:
            cell.fill = gray; cell.font = Font(bold=True)
            
        b_row = last_data_row + 2
        ws.cell(b_row, 1, "Кол-во проданных \nбилетов").font = Font(bold=True)
        ws.cell(b_row, 1).alignment = Alignment(wrap_text=True)
        
        cb = get_column_letter(2)
        clp = get_column_letter(last_col - 1)
        ci = get_column_letter(last_col)
        ws.cell(b_row, last_col).value = f"=SUM({cb}{b_row}:{clp}{b_row})"
        ws.cell(b_row, last_col).font = Font(bold=True)
        
        yellow = PatternFill("solid", fgColor="FFF2CC")
        for col in range(1, last_col + 1):
            ws.cell(b_row, col).fill = yellow
            
        rate_row = b_row + 2
        ws.cell(rate_row, 1, "Activation rate").font = Font(bold=True)
        for col in range(2, last_col):
            cl = get_column_letter(col)
            ws.cell(rate_row, col).value = f"={cl}{last_data_row}/{cl}{b_row}*100"
            ws.cell(rate_row, col).number_format = '0.0'
            
        green = PatternFill("solid", fgColor="D5F5E3")
        for col in range(1, last_col + 1):
            ws.cell(rate_row, col).fill = green
            
        avg_row = rate_row + 3
        ws.cell(avg_row, 1, "Activation rate\nсредняя").font = Font(bold=True)
        ws.cell(avg_row, 1).alignment = Alignment(wrap_text=True)
        ws.cell(avg_row, 2).value = f"=({ci}{last_data_row}-{cb}{last_data_row})/{ci}{b_row}"
        ws.cell(avg_row, 2).number_format = '0.00'
        
        for col_cells in ws.columns:
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(len(str(c.value or "")) for c in col_cells) + 3, 20)
        ws.column_dimensions["A"].width = 22

    return output_path
