import requests
import pandas as pd
from backend.reports.common import GRAFANA_URL, DATASOURCE_UID, HEADERS, get_time_boundaries
from openpyxl.styles import PatternFill, Font, Alignment

def generate_conversion(date_val: str, period_type: str, selected_parks: list, output_path: str):
    start_date, stop_date, _ = get_time_boundaries(date_val, period_type)
    
    def run_query(measurement, label):
        flux_query = f'''
        from(bucket: "Analytics_AvatarBD")
          |> range(start: {start_date}, stop: {stop_date})
          |> filter(fn: (r) => r["_measurement"] == "{measurement}")
          |> filter(fn: (r) => exists r["id"])
          |> group(columns: ["park", "id"])
          |> limit(n: 1)
          |> group(columns: ["park"])
          |> count(column: "_value")
          |> yield(name: "{label}")
        '''
        payload = {
            "queries": [{"refId": "A", "datasource": {"type": "influxdb", "uid": DATASOURCE_UID}, "query": flux_query}],
            "from": "0", "to": "9999999999999"
        }
        try:
            r = requests.post(f"{GRAFANA_URL}/api/ds/query", headers=HEADERS, json=payload, timeout=30)
            r.raise_for_status()
            results = {}
            for frame in r.json().get("results", {}).get("A", {}).get("frames", []):
                park_name = "Unknown"
                for field in frame.get("schema", {}).get("fields", []):
                    if "park" in field.get("labels", {}):
                        park_name = field["labels"]["park"]
                        break
                vals = frame.get("data", {}).get("values", [])
                if len(vals) > 0 and len(vals[0]) > 0:
                    results[park_name] = vals[0][0]
            return results
        except Exception as e:
            print(f"Error querying {measurement}: {e}")
            return {}

    avatars_data = run_query("AvatarLinkedInHome", "avatars")
    tasks_data = run_query("AchievementCounted", "tasks")

    all_parks = sorted([p for p in set(avatars_data.keys()).union(set(tasks_data.keys())) if p != "OFFICE" and p != "Unknown" and p in selected_parks])
    
    rows = []
    for park in all_parks:
        avatars = avatars_data.get(park, 0)
        tasks = tasks_data.get(park, 0)
        conversion = (tasks / avatars * 100) if avatars > 0 else 0
        rows.append({"Парк": park, "Создано аватаров": avatars, "Хотя бы одно задание": tasks, "Конверсия (%)": conversion})

    total_avatars = sum([avatars_data.get(p, 0) for p in all_parks])
    total_tasks = sum([tasks_data.get(p, 0) for p in all_parks])
    total_conversion = (total_tasks / total_avatars * 100) if total_avatars > 0 else 0
    rows.append({"Парк": "ИТОГО", "Создано аватаров": total_avatars, "Хотя бы одно задание": total_tasks, "Конверсия (%)": total_conversion})

    df = pd.DataFrame(rows)
    df_data = df.iloc[:-1].copy()
    df_data = df_data.sort_values("Парк")
    df_final = pd.concat([df_data, df.iloc[[-1]]], ignore_index=True)
    df_final["Конверсия (%)"] = df_final["Конверсия (%)"].round(1)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_final.to_excel(writer, sheet_name=f"Конверсия {date_val}", index=False)
        ws = writer.sheets[f"Конверсия {date_val}"]
        
        orange = PatternFill("solid", fgColor="FF6B00")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center = Alignment(horizontal="center")
        gray = PatternFill("solid", fgColor="E0E0E0")
        bold_font = Font(bold=True)
        green = PatternFill("solid", fgColor="D5F5E3")
        
        for col in range(1, len(df_final.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = orange; cell.font = header_font; cell.alignment = center
            
        last_row = ws.max_row
        for col in range(1, len(df_final.columns) + 1):
            cell = ws.cell(row=last_row, column=col)
            cell.fill = gray; cell.font = bold_font
            
        for row in range(2, last_row + 1):
            cell = ws.cell(row=row, column=4)
            cell.fill = green; cell.number_format = '0.0"%"'
            
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 16

    return output_path
