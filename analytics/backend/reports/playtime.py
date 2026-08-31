import requests
import pandas as pd
from datetime import datetime
from backend.reports.common import GRAFANA_URL, DATASOURCE_UID, HEADERS, get_time_boundaries
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def generate_playtime(date_val: str, period_type: str, selected_parks: list, output_path: str, strict_mode: bool = False):
    start_date, stop_date, _ = get_time_boundaries(date_val, period_type)
    
    # First query: actual playtime from SessionEnd
    flux_query = f'''
    from(bucket: "Analytics_AvatarBD")
      |> range(start: {start_date}, stop: {stop_date})
      |> filter(fn: (r) => r["_measurement"] == "SessionEnd")
      |> filter(fn: (r) => r["_field"] == "SessionDuration")
      |> filter(fn: (r) => r["zone"] >= "hp" and r["zone"] < "hq")
      |> map(fn: (r) => ({{ r with _value: float(v: r._value) }}))
      |> group(columns: ["park", "zone", "scene"])
      |> sum(column: "_value")
      |> yield(name: "playtime_stats")
    '''
    
    # Second query: just to discover other games that don't have SessionEnd
    flux_query_others = f'''
    from(bucket: "Analytics_AvatarBD")
      |> range(start: {start_date}, stop: {stop_date})
      |> filter(fn: (r) => r["_measurement"] == "PlayerEnteredInstallation")
      |> filter(fn: (r) => r["_field"] == "value")
      |> filter(fn: (r) => r["zone"] >= "hp" and r["zone"] < "hq")
      |> group(columns: ["park", "zone", "scene"])
      |> count(column: "_value")
      |> yield(name: "other_games")
    '''

    payload = {
        "queries": [
            {"refId": "A", "datasource": {"type": "influxdb", "uid": DATASOURCE_UID}, "query": flux_query, "hide": False},
            {"refId": "B", "datasource": {"type": "influxdb", "uid": DATASOURCE_UID}, "query": flux_query_others, "hide": False}
        ],
        "from": "0", "to": "9999999999999"
    }

    response = requests.post(f"{GRAFANA_URL}/api/ds/query", headers=HEADERS, json=payload, timeout=30)
    response.raise_for_status()
    
    results_dict = {}
    
    # Process SessionEnd (playtime)
    frames_a = response.json().get("results", {}).get("A", {}).get("frames", [])
    for frame in frames_a:
        schema = frame.get("schema", {})
        labels = {}
        val_idx = -1
        fields = schema.get("fields", [])
        for i, field in enumerate(fields):
            if "labels" in field:
                labels = field["labels"]
            if field.get("name") == "_value":
                val_idx = i
                
        if val_idx == -1 and len(fields) > 0:
            val_idx = len(fields) - 1
                
        park = labels.get("park", "Unknown")
        if park not in selected_parks:
            continue
            
        zone = labels.get("zone", "Unknown")
        if zone in ["hp-quest-portal", "hp-avatars-hologram"]:
            continue

        scene = labels.get("scene", "Unknown")
        
        vals = frame.get("data", {}).get("values", [])
        if val_idx >= 0 and len(vals) > val_idx and len(vals[val_idx]) > 0:
            val_seconds = vals[val_idx][0]
            if val_seconds is not None:
                val_hours = val_seconds / 3600.0
                results_dict[(park, zone, scene)] = val_hours
                
    # Process PlayerEnteredInstallation (others with 0 hours)
    if not strict_mode:
        frames_b = response.json().get("results", {}).get("B", {}).get("frames", [])
        for frame in frames_b:
            schema = frame.get("schema", {})
            labels = {}
            fields = schema.get("fields", [])
            for i, field in enumerate(fields):
                if "labels" in field:
                    labels = field["labels"]
                    
            park = labels.get("park", "Unknown")
            if park not in selected_parks:
                continue
                
            zone = labels.get("zone", "Unknown")
            if zone in ["hp-quest-portal", "hp-avatars-hologram"]:
                continue

            scene = labels.get("scene", "Unknown")
            
            if (park, zone, scene) not in results_dict:
                results_dict[(park, zone, scene)] = 0.0

    results = []
    for (park, zone, scene), hours in results_dict.items():
        results.append({
            "Парк": park,
            "Игра (zone)": zone,
            "Тема (scene)": scene,
            "Часов": hours
        })

    if not results:
        raise Exception("Нет данных за выбранный период")

    df = pd.DataFrame(results)
    
    # 1. Сводная таблица по паркам, темам и играм
    pivot_df = df.pivot_table(
        index=["Парк", "Игра (zone)", "Тема (scene)"], 
        values="Часов", 
        aggfunc="sum"
    ).fillna(0)
    
    # 2. Общий рейтинг игр (сумма по всем паркам)
    game_rating = df.pivot_table(
        index=["Игра (zone)", "Тема (scene)"],
        values="Часов",
        aggfunc="sum"
    ).sort_values(by="Часов", ascending=False)
    
    # Запись в Excel
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pivot_df.to_excel(writer, sheet_name="По паркам")
        game_rating.to_excel(writer, sheet_name="Рейтинг игр")
        
        # Форматирование листов
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            # Заголовки (оранжевые)
            orange = PatternFill("solid", fgColor="FF6B00")
            for cell in ws[1]:
                cell.fill = orange
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            
            # Округляем до 2 знаков в колонках со значениями часов
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'
                        
            # Авто-ширина
            for col_cells in ws.columns:
                max_length = 0
                for cell in col_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width = adjusted_width

    return output_path
