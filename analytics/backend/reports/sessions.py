import requests
import pandas as pd
from backend.reports.common import GRAFANA_URL, DATASOURCE_UID, get_headers, get_time_boundaries
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def generate_sessions(date_val: str, period_type: str, selected_parks: list, output_path: str, strict_mode: bool = False):
    start_date, stop_date, _ = get_time_boundaries(date_val, period_type)
    
    # We group by park, zone, scene, and _measurement to handle games that do not have SessionStart.
    flux_query = f'''
    from(bucket: "Analytics_AvatarBD")
      |> range(start: {start_date}, stop: {stop_date})
      |> filter(fn: (r) => r["_measurement"] == "SessionStart" or r["_measurement"] == "PlayerEnteredInstallation")
      |> filter(fn: (r) => r["_field"] == "value")
      |> filter(fn: (r) => r["zone"] >= "hp" and r["zone"] < "hq")
      |> group(columns: ["park", "zone", "scene", "_measurement"])
      |> count(column: "_value")
      |> yield(name: "sessions_stats")
    '''

    payload = {
        "queries": [{"refId": "A", "datasource": {"type": "influxdb", "uid": DATASOURCE_UID}, "query": flux_query, "hide": False}],
        "from": "0", "to": "9999999999999"
    }

    response = requests.post(f"{GRAFANA_URL}/api/ds/query", headers=get_headers(), json=payload, timeout=30)

    response.raise_for_status()
    
    frames = response.json().get("results", {}).get("A", {}).get("frames", [])
    
    # Store counts in a dictionary: (park, zone, scene) -> {"SessionStart": count, "PlayerEnteredInstallation": count}
    raw_results = {}
    
    for frame in frames:
        schema = frame.get("schema", {})
        labels = {}
        val_idx = -1
        fields = schema.get("fields", [])
        for i, field in enumerate(fields):
            if "labels" in field:
                labels = field["labels"]
            if field.get("name") == "_value":
                val_idx = i
                
        if val_idx == -1 and len(fields) > 0:
            val_idx = len(fields) - 1
                
        park = labels.get("park", "Unknown")
        if park not in selected_parks:
            continue
            
        zone = labels.get("zone", "Unknown")
        if zone in ["hp-quest-portal", "hp-avatars-hologram"]:
            continue

        scene = labels.get("scene", "Unknown")
        measurement = schema.get("name", "Unknown") # or extract from labels if it is there
        if "_measurement" in labels:
            measurement = labels["_measurement"]
            
        vals = frame.get("data", {}).get("values", [])
        if val_idx >= 0 and len(vals) > val_idx and len(vals[val_idx]) > 0:
            val_count = vals[val_idx][0]
            if val_count is not None:
                key = (park, zone, scene)
                if key not in raw_results:
                    raw_results[key] = {}
                raw_results[key][measurement] = val_count

    results = []
    for (park, zone, scene), meas_counts in raw_results.items():
        if strict_mode:
            count = meas_counts.get("SessionStart", 0)
        else:
            # Prefer SessionStart, fallback to PlayerEnteredInstallation
            count = meas_counts.get("SessionStart", meas_counts.get("PlayerEnteredInstallation", 0))
            
        if count > 0:
            results.append({
                "Парк": park,
                "Игра (zone)": zone,
                "Тема (scene)": scene,
                "Сессий": count
            })

    if not results:
        raise Exception("Нет данных за выбранный период")

    df = pd.DataFrame(results)
    
    # 1. Сводная таблица по паркам, темам и играм
    pivot_df = df.pivot_table(
        index=["Парк", "Игра (zone)", "Тема (scene)"], 
        values="Сессий", 
        aggfunc="sum"
    ).fillna(0).astype(int)
    
    # 2. Общий рейтинг игр (сумма по всем паркам)
    game_rating = df.pivot_table(
        index=["Игра (zone)", "Тема (scene)"],
        values="Сессий",
        aggfunc="sum"
    ).sort_values(by="Сессий", ascending=False).astype(int)
    
    # Запись в Excel
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pivot_df.to_excel(writer, sheet_name="По паркам")
        game_rating.to_excel(writer, sheet_name="Рейтинг игр")
        
        # Форматирование
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            orange = PatternFill("solid", fgColor="FF6B00")
            for cell in ws[1]:
                cell.fill = orange
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
                        
            # Авто-ширина
            for col_cells in ws.columns:
                max_length = 0
                for cell in col_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width = (max_length + 2)

    return output_path
